

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""file_size.py

Scan a folder for PDF files (up to a chosen directory depth), compute file sizes,
then export results to an Excel (.xlsx) report.

Works on macOS/Windows/Linux.

Usage:
  - Run directly: python3 file_size.py
  - Or make executable and double-click / run from terminal.

Notes on depth:
  - depth = 0  -> only the selected folder
  - depth = 1  -> include immediate subfolders
  - depth = 2  -> include subfolders two levels deep
"""

from __future__ import annotations

import os
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


@dataclass
class ScanResult:
    index: int
    rel_dir: str
    filename: str
    size_bytes: int
    size_mb: float
    full_path: str
    mtime: str


def _human_mb(size_bytes: int) -> float:
    return round(size_bytes / (1024 * 1024), 3)


def _safe_relpath(child: Path, root: Path) -> str:
    try:
        rel = child.relative_to(root)
        return str(rel)
    except Exception:
        return str(child)


def iter_pdf_files(root: Path, depth: int) -> Iterable[Path]:
    """Yield PDF file Paths under root, limited by directory depth."""
    root = root.resolve()
    if depth < 0:
        depth = 0

    # Walk manually to control depth
    # current_depth: root is 0
    stack: List[Tuple[Path, int]] = [(root, 0)]

    while stack:
        folder, d = stack.pop()
        try:
            with os.scandir(folder) as it:
                for entry in it:
                    # Skip hidden folders/files by default? We keep them, but you can uncomment below.
                    # if entry.name.startswith('.'):
                    #     continue

                    try:
                        if entry.is_file(follow_symlinks=False):
                            if entry.name.lower().endswith(".pdf"):
                                yield Path(entry.path)
                        elif entry.is_dir(follow_symlinks=False):
                            if d < depth:
                                stack.append((Path(entry.path), d + 1))
                    except PermissionError:
                        # Skip unreadable entries
                        continue
        except PermissionError:
            continue


def scan_pdfs(root: Path, depth: int) -> List[ScanResult]:
    results: List[ScanResult] = []
    idx = 0
    for p in iter_pdf_files(root, depth):
        try:
            st = p.stat()
            size_b = int(st.st_size)
            mtime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(st.st_mtime))
        except Exception:
            # If stat fails, skip
            continue

        idx += 1
        rel = _safe_relpath(p.parent, root)
        results.append(
            ScanResult(
                index=idx,
                rel_dir=rel if rel != "." else "",
                filename=p.name,
                size_bytes=size_b,
                size_mb=_human_mb(size_b),
                full_path=str(p),
                mtime=mtime,
            )
        )

    # Sort: bigger first, then name
    results.sort(key=lambda r: (-r.size_bytes, r.filename.lower()))

    # Re-number after sorting
    for i, r in enumerate(results, start=1):
        r.index = i

    return results


def export_to_excel(
    out_path: Path,
    root: Path,
    depth: int,
    results: List[ScanResult],
) -> None:
    wb = Workbook()

    # Sheet 1: details
    ws = wb.active
    ws.title = "PDF大小明细"

    header = [
        "序号",
        "相对目录",
        "文件名",
        "大小(Bytes)",
        "大小(MB)",
        "修改时间",
        "完整路径",
    ]
    ws.append(header)

    for hcell in ws[1]:
        hcell.font = Font(bold=True)
        hcell.alignment = Alignment(horizontal="center", vertical="center")

    for r in results:
        ws.append(
            [
                r.index,
                r.rel_dir,
                r.filename,
                r.size_bytes,
                r.size_mb,
                r.mtime,
                r.full_path,
            ]
        )

    # Formatting
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"

    # Column widths
    col_widths = {
        1: 7,
        2: 26,
        3: 42,
        4: 14,
        5: 12,
        6: 20,
        7: 70,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Align numeric columns
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=7):
        row[0].alignment = Alignment(horizontal="center")
        row[3].alignment = Alignment(horizontal="right")
        row[4].alignment = Alignment(horizontal="right")

    # Sheet 2: summary
    ws2 = wb.create_sheet("汇总")
    ws2.append(["扫描根目录", str(root)])
    ws2.append(["扫描目录级数(depth)", depth])
    ws2.append(["PDF数量", len(results)])
    total_bytes = sum(r.size_bytes for r in results)
    ws2.append(["总大小(Bytes)", total_bytes])
    ws2.append(["总大小(MB)", round(total_bytes / (1024 * 1024), 3)])

    ws2.append([])
    ws2.append(["Top 20 最大PDF"]) 
    ws2.append(["序号", "文件名", "大小(MB)", "相对目录"]) 

    topn = results[:20]
    for i, r in enumerate(topn, start=1):
        ws2.append([i, r.filename, r.size_mb, r.rel_dir])

    for hcell in ws2[8]:
        hcell.font = Font(bold=True)
        hcell.alignment = Alignment(horizontal="center", vertical="center")

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 30

    wb.save(out_path)


class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("PDF 文件大小检查")
        self.geometry("900x620")
        self.minsize(820, 560)

        self.root_dir = tk.StringVar(value="")
        self.depth = tk.IntVar(value=2)
        self.include_zero = tk.BooleanVar(value=True)

        self._results: List[ScanResult] = []

        self._build_ui()

    def _build_ui(self) -> None:
        # Top controls
        frm = ttk.Frame(self, padding=12)
        frm.pack(fill="x")

        ttk.Label(frm, text="扫描目录:").grid(row=0, column=0, sticky="w")
        entry = ttk.Entry(frm, textvariable=self.root_dir)
        entry.grid(row=0, column=1, sticky="ew", padx=(8, 8))

        ttk.Button(frm, text="选择…", command=self.pick_dir).grid(row=0, column=2, sticky="e")

        ttk.Label(frm, text="目录级数(depth):").grid(row=1, column=0, sticky="w", pady=(10, 0))
        sp = ttk.Spinbox(frm, from_=0, to=20, textvariable=self.depth, width=6)
        sp.grid(row=1, column=1, sticky="w", padx=(8, 0), pady=(10, 0))

        ttk.Label(
            frm,
            text="说明：0=仅当前目录，1=含下一层子目录，2=含两层…",
            foreground="#555",
        ).grid(row=1, column=1, sticky="w", padx=(110, 0), pady=(10, 0))

        btns = ttk.Frame(frm)
        btns.grid(row=0, column=3, rowspan=2, sticky="e", padx=(12, 0))
        ttk.Button(btns, text="开始扫描", command=self.run_scan).pack(fill="x")
        ttk.Button(btns, text="导出Excel", command=self.export_excel).pack(fill="x", pady=(8, 0))

        frm.columnconfigure(1, weight=1)

        # Progress + status
        sfrm = ttk.Frame(self, padding=(12, 0, 12, 8))
        sfrm.pack(fill="x")
        self.progress = ttk.Progressbar(sfrm, mode="indeterminate")
        self.progress.pack(fill="x")
        self.status = tk.StringVar(value="请选择扫描目录，然后点击开始扫描")
        ttk.Label(sfrm, textvariable=self.status).pack(anchor="w", pady=(6, 0))

        # Table
        tfrm = ttk.Frame(self, padding=(12, 0, 12, 12))
        tfrm.pack(fill="both", expand=True)

        cols = ("序号", "相对目录", "文件名", "大小(MB)", "大小(Bytes)", "修改时间")
        self.tree = ttk.Treeview(tfrm, columns=cols, show="headings", height=18)
        for c in cols:
            self.tree.heading(c, text=c)

        self.tree.column("序号", width=60, anchor="center")
        self.tree.column("相对目录", width=220, anchor="w")
        self.tree.column("文件名", width=380, anchor="w")
        self.tree.column("大小(MB)", width=100, anchor="e")
        self.tree.column("大小(Bytes)", width=120, anchor="e")
        self.tree.column("修改时间", width=150, anchor="center")

        yscroll = ttk.Scrollbar(tfrm, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=yscroll.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")

        tfrm.rowconfigure(0, weight=1)
        tfrm.columnconfigure(0, weight=1)

        # Bottom hint
        hint = ttk.Label(
            self,
            padding=(12, 0, 12, 12),
            text="提示：双击表格行可复制完整路径到剪贴板（方便定位文件）。",
            foreground="#555",
        )
        hint.pack(anchor="w")

        self.tree.bind("<Double-1>", self.copy_selected_path)

    def pick_dir(self) -> None:
        d = filedialog.askdirectory(title="选择要扫描的目录")
        if d:
            self.root_dir.set(d)

    def _clear_table(self) -> None:
        for item in self.tree.get_children():
            self.tree.delete(item)

    def _fill_table(self, results: List[ScanResult]) -> None:
        self._clear_table()
        for r in results:
            self.tree.insert(
                "",
                "end",
                values=(r.index, r.rel_dir, r.filename, f"{r.size_mb:.3f}", r.size_bytes, r.mtime),
            )

    def run_scan(self) -> None:
        root = self.root_dir.get().strip()
        if not root:
            messagebox.showwarning("缺少目录", "请先选择要扫描的目录。")
            return

        root_path = Path(root)
        if not root_path.exists() or not root_path.is_dir():
            messagebox.showerror("目录无效", "所选目录不存在或不是文件夹。")
            return

        depth = int(self.depth.get())

        self.status.set("正在扫描PDF…")
        self.progress.start(10)
        self.update_idletasks()

        try:
            results = scan_pdfs(root_path, depth)
        except Exception as e:
            self.progress.stop()
            self.status.set("扫描失败")
            messagebox.showerror("扫描失败", f"扫描过程中发生错误：\n{e}")
            return
        finally:
            self.progress.stop()

        self._results = results
        self._fill_table(results)

        total_bytes = sum(r.size_bytes for r in results)
        self.status.set(
            f"扫描完成：PDF数量 {len(results)}，总大小 {_human_mb(total_bytes):.3f} MB（按大小从大到小排序）"
        )

    def export_excel(self) -> None:
        if not self._results:
            messagebox.showinfo("暂无数据", "请先开始扫描，生成结果后再导出。")
            return

        root = Path(self.root_dir.get().strip()).resolve()
        suggested = f"PDF大小报告_{root.name}.xlsx"

        out = filedialog.asksaveasfilename(
            title="保存Excel报告",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile=suggested,
        )
        if not out:
            return

        out_path = Path(out)

        try:
            export_to_excel(out_path, root, int(self.depth.get()), self._results)
        except Exception as e:
            messagebox.showerror("导出失败", f"导出Excel失败：\n{e}")
            return

        messagebox.showinfo("导出成功", f"已导出：\n{out_path}")

    def copy_selected_path(self, _event=None) -> None:
        sel = self.tree.selection()
        if not sel:
            return

        item = sel[0]
        values = self.tree.item(item, "values")
        # values: index, rel_dir, filename, size_mb, size_bytes, mtime
        if len(values) < 3:
            return

        rel_dir = values[1]
        filename = values[2]
        root = Path(self.root_dir.get().strip())
        full = (root / rel_dir / filename).resolve() if rel_dir else (root / filename).resolve()

        self.clipboard_clear()
        self.clipboard_append(str(full))
        self.status.set(f"已复制路径：{full}")


def _use_ttk_theme(root: tk.Tk) -> None:
    # Prefer a modern-looking theme if available
    style = ttk.Style(root)
    # Some platforms have: 'aqua' (mac), 'vista' (win), 'clam' (cross)
    for candidate in ("aqua", "vista", "clam"):
        try:
            style.theme_use(candidate)
            break
        except tk.TclError:
            continue
def run(folder: str, scan_depth: int, out_xlsx: str | None = None) -> int:
    """
    folder: 要扫描的目录
    scan_depth: 扫描层级（0=不限递归，1=当前，2=下一层...）
    out_xlsx: 输出excel路径（可选）
    return: 0=成功，非0=失败
    """
    # TODO: 把你 main() 里面真正干活的逻辑搬到这里
    # 举例（你按自己现有逻辑替换）：
    # results = scan_folder(folder, scan_depth)
    # export_excel(results, out_xlsx or default_path)
    return 0

def main() -> int:
    app = App()
    _use_ttk_theme(app)
    app.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
